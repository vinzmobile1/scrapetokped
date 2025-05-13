import streamlit as st
import requests
import json
import uuid
from urllib.parse import urlparse, parse_qs, unquote
import pandas as pd
import time
import io
# from io import BytesIO # BytesIO sudah diimpor dari io
from openpyxl import Workbook # Import Workbook jika belum ada
from openpyxl.utils import get_column_letter # Untuk mendapatkan huruf kolom
from openpyxl.styles import Font # Untuk style hyperlink jika diinginkan

# === Helper Function ===
def get_nested_value(data_dict, keys, default=None):
    current = data_dict
    for key in keys:
        if isinstance(current, dict) and key in current:
            current = current[key]
        elif isinstance(current, list) and isinstance(key, int) and 0 <= key < len(current):
            current = current[key]
        else:
            return default
    return current

# === Fungsi Helper untuk Format Durasi Dinamis ===
def format_duration(seconds):
    if seconds < 0:
        return "segera"
    if seconds < 60:
        return f"{seconds:.1f} detik"
    elif seconds < 3600:
        minutes = seconds / 60
        return f"{minutes:.1f} menit"
    else:
        hours = seconds / 3600
        return f"{hours:.1f} jam"

# === Step 1: Fetch All Product URLs ===
# (Fungsi fetch_all_product_urls_from_shop tidak berubah, jadi saya singkat di sini)
def fetch_all_product_urls_from_shop(headers, sid):
    # ... (kode fungsi fetch_all_product_urls_from_shop tetap sama) ...
    url = "https://gql.tokopedia.com/graphql/ShopProducts"

    def get_payload(page):
        return [{
            "operationName": "ShopProducts",
            "variables": {
                "source": "shop",
                "sid": sid,
                "page": page,
                "perPage": 80,
                "etalaseId": "etalase",
                "sort": 1,
                "user_districtId": "2274",
                "user_cityId": "176",
                "user_lat": "0",
                "user_long": "0"
            },
            "query": """
            query ShopProducts($sid: String!, $source: String, $page: Int, $perPage: Int, $keyword: String, $etalaseId: String, $sort: Int, $user_districtId: String, $user_cityId: String, $user_lat: String, $user_long: String) {
                GetShopProduct(shopID: $sid, source: $source, filter: {
                    page: $page, perPage: $perPage, fkeyword: $keyword,
                    fmenu: $etalaseId, sort: $sort,
                    user_districtId: $user_districtId,
                    user_cityId: $user_cityId,
                    user_lat: $user_lat, user_long: $user_long
                }) {
                    links { next }
                    data { product_url }
                }
            }
            """
        }]

    product_urls = []
    page = 1
    st.write("Memulai pengambilan URL produk...")
    page_progress_text = st.empty() # Placeholder for page progress
    start_fetch_urls_time = time.time()
    page_count = 0
    while True:
        page_count += 1
        elapsed_fetch_urls_time = time.time() - start_fetch_urls_time
        page_progress_text.text(f"Mengambil URL dari halaman {page_count}... Waktu berjalan: {format_duration(elapsed_fetch_urls_time)}")

        response = requests.post(url, headers=headers, json=get_payload(page))
        if response.status_code != 200:
            st.error(f"Gagal mengambil URL dari halaman {page}. Status: {response.status_code}")
            st.json(response.text) # Show error response
            break

        try:
            result_data = response.json()
             # Lebih aman dalam mengakses data
            get_shop_product_data = get_nested_value(result_data, [0, 'data', 'GetShopProduct'])
            if not get_shop_product_data:
                 st.error(f"Data 'GetShopProduct' tidak ditemukan atau format tidak sesuai di halaman {page}.")
                 st.json(result_data)
                 break
            urls_on_page = [p['product_url'] for p in get_shop_product_data.get('data', []) if p and 'product_url' in p]
            product_urls.extend(urls_on_page)
        except (IndexError, KeyError, TypeError, json.JSONDecodeError) as e:
            st.error(f"Error parsing JSON dari halaman {page}: {e}")
            try:
                st.json(response.text) # Show problematic JSON text if possible
            except:
                st.text("(Gagal menampilkan detail response error)")
            break

        # Cek 'links' dan 'next' dengan lebih aman
        links_data = get_nested_value(result_data, [0, 'data', 'GetShopProduct', 'links'])
        if not links_data or not links_data.get('next'):
             break
        page += 1
        time.sleep(0.5) # Jeda sedikit

    total_fetch_urls_time = time.time() - start_fetch_urls_time
    page_progress_text.text(f"Selesai mengambil URL. Total {len(product_urls)} URL ditemukan dalam {format_duration(total_fetch_urls_time)}.")
    return product_urls


# === Step 2: Fetch Product Detail ===
# (Fungsi fetch_tokopedia_product_data tidak berubah, jadi saya singkat di sini)
def fetch_tokopedia_product_data(product_url, headers_template, show_logs=False):
    # ... (kode fungsi fetch_tokopedia_product_data tetap sama) ...
    try:
        parsed_url = urlparse(product_url)
        path_parts = [part for part in parsed_url.path.split('/') if part]
        if len(path_parts) < 2:
            if show_logs: st.warning(f"URL tidak valid (path parts < 2): {product_url}")
            return None

        # Ambil 2 bagian terakhir path sebagai shop_domain dan product_key (lebih robust)
        shop_domain_val, product_key_val = path_parts[-2], path_parts[-1]
        ext_param_val = parse_qs(parsed_url.query).get('extParam', [''])[0]
    except Exception as e:
        if show_logs: st.warning(f"Error parsing URL {product_url}: {e}")
        return None

    request_url = "https://gql.tokopedia.com/graphql/PDPGetLayoutQuery"
    graphql_query = """
    query PDPGetLayoutQuery($shopDomain: String, $productKey: String, $layoutID: String, $apiVersion: Float, $userLocation: pdpUserLocation, $extParam: String, $tokonow: pdpTokoNow, $deviceID: String) {
      pdpGetLayout(shopDomain: $shopDomain, productKey: $productKey, layoutID: $layoutID, apiVersion: $apiVersion, userLocation: $userLocation, extParam: $extParam, tokonow: $tokonow, deviceID: $deviceID) {
        requestID name pdpSession basicInfo {
          id: productID shopID shopName txStats { countSold } stats { countReview rating } # ttsPID mungkin tidak di sini
        }
         # Coba cari ttsPID di level atas atau di components jika ada
         # ttsPID # Jika ada di level ini
         # components { ... } # Cari di dalam components jika perlu
      }
    }
    """

    payload = [{
        "operationName": "PDPGetLayoutQuery",
        "variables": {
            "shopDomain": shop_domain_val, # Gunakan shop_domain dari URL
            "productKey": product_key_val,
            "layoutID": "",
            "apiVersion": 1.1, # Coba versi sedikit lebih baru
            "tokonow": {"shopID": "", "whID": "0", "serviceType": ""},
            "deviceID": str(uuid.uuid4()),
            "userLocation": {"cityID": "176", "addressID": "", "districtID": "2274", "postalCode": "", "latlon": ""},
            "extParam": ext_param_val
        },
        "query": graphql_query
    }]

    try:
        if show_logs:
            st.write(f"--- MENGIRIM REQUEST UNTUK: {product_url}")
        response = requests.post(request_url, json=payload, headers=headers_template, timeout=30)
        response.raise_for_status()
        response_json = response.json()

        # Cek error GraphQL
        if "errors" in response_json[0] and response_json[0]["errors"]:
            if show_logs: st.error(f"GraphQL Error untuk {product_url}: {response_json[0]['errors']}")
            return None

        pdp_layout_data = get_nested_value(response_json, [0, "data", "pdpGetLayout"])
        if not pdp_layout_data:
             if show_logs: st.warning(f"Data 'pdpGetLayout' kosong untuk {product_url}")
             # st.json(response_json) # Tampilkan response jika data kosong
             return None

        return pdp_layout_data

    except requests.exceptions.RequestException as e:
        if show_logs: st.error(f"Request gagal untuk {product_url}: {e}")
        return None
    except (IndexError, KeyError, json.JSONDecodeError) as e:
        if show_logs: st.error(f"Error parsing JSON response untuk {product_url}: {e}")
        try:
            if show_logs: st.text(response.text) # Tampilkan raw text jika JSON error
        except: pass
        return None
    except Exception as e: # Tangkap error lainnya
        if show_logs: st.error(f"Error tidak terduga saat memproses {product_url}: {e}")
        return None


# === Step 3: Extract Data ===
# (Fungsi extract_product_details tidak berubah, jadi saya singkat di sini)
def extract_product_details(pdp_data):
    # ... (kode fungsi extract_product_details tetap sama) ...
    if not pdp_data:
        return None

    # Ekstraksi dasar
    basic_info = pdp_data.get('basicInfo', {})
    stats = basic_info.get('stats', {})
    tx_stats = basic_info.get('txStats', {})

    details = {
        'ProductID': basic_info.get('productID'), # Ganti 'id' ke 'productID'
        'ShopID': basic_info.get('shopID'),
        'ShopName': basic_info.get('shopName'),
        'CountSold': tx_stats.get('countSold'),
        'CountReview': stats.get('countReview'),
        'Rating': stats.get('rating'),
        'ProductName': pdp_data.get('name'), # Default nama
        'PriceValue': None,
        'ttsPID': None # Default None, lokasi ttsPID perlu diverifikasi lagi di response API
        # Coba cari ttsPID di pdp_data langsung jika ada key-nya
        # 'ttsPID': pdp_data.get('ttsPID'), # Uncomment jika ttsPID ada di level ini
    }

    # Ambil dari pdpSession jika ada (seringkali lebih akurat untuk nama/harga)
    pdp_session_str = pdp_data.get('pdpSession')
    if pdp_session_str:
        try:
            session_data = json.loads(pdp_session_str)
            # Override nama jika ada ('pn' atau 'productName')
            if 'pn' in session_data: details['ProductName'] = session_data['pn']
            elif 'productName' in session_data: details['ProductName'] = session_data['productName']
            # Ambil harga jika ada ('pr' atau 'price')
            if 'pr' in session_data: details['PriceValue'] = session_data['pr']
            elif 'price' in session_data: details['PriceValue'] = session_data['price']
            # Coba cari ttsPID di session ('tid'?) jika belum ketemu
            if details['ttsPID'] is None: details['ttsPID'] = session_data.get('tid')

        except (json.JSONDecodeError, TypeError):
             # Biarkan fallback name & price jika session error
             if details['ProductName'] is None: details['ProductName'] = basic_info.get('name') # Fallback lagi jika perlu
             pass
        except Exception as e:
            st.warning(f"Error processing pdpSession: {e}")

    # Final fallback name jika masih kosong
    if not details['ProductName']:
         details['ProductName'] = basic_info.get('name')

    return details

# === Streamlit UI ===
st.set_page_config(layout="wide")
st.title("Tokopedia Produk Scraper")
sid_input = st.text_input("Masukkan SID toko Tokopedia:", "14799089")
show_logs = st.checkbox("Tampilkan log detail proses (memperlambat UI)")

if st.button("Execute", key="execute_button"):
    if not sid_input:
        st.error("SID Toko tidak boleh kosong!")
    elif not sid_input.isdigit():
        st.error("SID Toko harus berupa angka!")
    else:
        st.info("Memulai proses scraping...")

        headers = {
            'sec-ch-ua-platform': '"Windows"',
            'x-version': '7d93f84',
            'sec-ch-ua': '"Chromium";v="120", "Google Chrome";v="120", "Not_A Brand";v="8"', # Sesuaikan dgn browser Anda
            'x-price-center': 'true',
            'sec-ch-ua-mobile': '?0',
            'x-source': 'tokopedia-lite',
            'x-tkpd-akamai': 'pdpGetLayout', # Mungkin perlu diubah/dihapus
            'x-device': 'desktop',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36', # Sesuaikan
            'accept': 'application/json', # Biasanya API GQL menerima ini
            'content-type': 'application/json',
            'x-tkpd-lite-service': 'zeus', # Header spesifik Tokopedia
            'Origin': 'https://www.tokopedia.com',
            'Referer': 'https://www.tokopedia.com/',
            'Accept-Language': 'en-US,en;q=0.9,id;q=0.8',
            # Header lain jika perlu (misal: sec-fetch-*, dsb)
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-site',
        }


        urls = fetch_all_product_urls_from_shop(headers, sid_input)

        if not urls:
            st.warning("Tidak ada URL produk yang ditemukan atau gagal mengambil URL.")
        else:
            st.info(f"Mengambil detail untuk {len(urls)} produk...")
            all_data = []

            progress_bar = st.progress(0)
            status_text = st.empty()
            start_time = time.time()
            processed_count = 0

            for i, url_item in enumerate(urls):
                pdp = fetch_tokopedia_product_data(url_item, headers, show_logs=show_logs)
                if pdp:
                    data = extract_product_details(pdp)
                    if data and data.get('ProductID'): # Perlu ProductID sebagai minimal data valid
                        data['ProductURL'] = url_item
                        all_data.append(data)
                        processed_count += 1
                    elif show_logs:
                        st.warning(f"Data tidak lengkap atau ProductID kosong untuk URL: {url_item}")

                # Update progress
                progress_percentage = (i + 1) / len(urls)
                elapsed_time = time.time() - start_time

                eta_str = ""
                if i > 5 and elapsed_time > 1: # Estimasi setelah beberapa data
                    time_per_item = elapsed_time / (i + 1)
                    remaining_items = len(urls) - (i + 1)
                    eta = time_per_item * remaining_items
                    eta_str = f" | ETA: {format_duration(eta)}"
                else:
                    eta_str = ""

                progress_bar.progress(progress_percentage)
                status_text.text(f"Memproses {i+1}/{len(urls)} produk... ({processed_count} berhasil) | Waktu berjalan: {format_duration(elapsed_time)}{eta_str}")

                time.sleep(0.15) # Jeda antar detail produk, bisa disesuaikan

            total_elapsed_time = time.time() - start_time
            status_text.text(f"Selesai! Total waktu pemrosesan: {format_duration(total_elapsed_time)}. Berhasil memproses {processed_count} dari {len(urls)} produk.")

            if all_data:
                df_raw = pd.DataFrame(all_data)

                # --- Persiapan Tipe Data DataFrame sebelum Export ---
                # Kolom yang diharapkan sebagai teks
                text_cols = ['ShopID', 'ShopName', 'ProductID', 'ttsPID', 'ProductName', 'ProductURL']
                for col in text_cols:
                    if col in df_raw.columns:
                        # Pastikan kolom ada dan konversi ke string, isi NaN dengan string kosong
                        df_raw[col] = df_raw[col].astype(str).fillna('')

                # Kolom yang diharapkan sebagai integer
                int_cols = ['PriceValue', 'CountSold', 'CountReview']
                for col in int_cols:
                     if col in df_raw.columns:
                        # Konversi ke numerik (paksa error jadi NaN), isi NaN dgn 0, lalu konversi ke Int64 (nullable integer)
                        df_raw[col] = pd.to_numeric(df_raw[col], errors='coerce').fillna(0).astype('Int64')

                # Kolom yang diharapkan sebagai float
                float_cols = ['Rating']
                for col in float_cols:
                     if col in df_raw.columns:
                        # Konversi ke numerik (paksa error jadi NaN), biarkan NaN jika ada (atau isi dgn 0.0 jika perlu)
                        df_raw[col] = pd.to_numeric(df_raw[col], errors='coerce').astype(float) # Bisa juga .fillna(0.0) jika mau


                # --- Urutan Kolom ---
                desired_column_order = [
                    'ShopID', 'ShopName', 'ProductID', 'ttsPID', 'ProductName',
                    'PriceValue', 'CountSold', 'CountReview', 'Rating', 'ProductURL'
                ]
                current_columns = [col for col in desired_column_order if col in df_raw.columns]
                for col in df_raw.columns:
                    if col not in current_columns:
                        current_columns.append(col) # Tambahkan kolom ekstra di akhir

                df = df_raw[current_columns]

                st.success(f"Berhasil mengambil {len(df)} produk dari {len(urls)} URL yang diproses.")
                st.dataframe(df)

                # --- Convert to Excel with Formatting ---
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Produk')

                    # Dapatkan workbook dan worksheet object setelah data ditulis
                    workbook = writer.book
                    worksheet = writer.sheets['Produk']

                    # --- Terapkan Formatting ---
                    # (Mulai dari baris 2 karena baris 1 adalah header)
                    max_row = worksheet.max_row
                    for col_idx_0based, col_name in enumerate(df.columns):
                        col_letter = get_column_letter(col_idx_0based + 1) # +1 karena openpyxl 1-based index
                        format_code = None # Default

                        # Tentukan format berdasarkan nama kolom
                        if col_name in ['ShopID', 'ShopName', 'ProductID', 'ttsPID', 'ProductName']:
                            format_code = '@' # Format Text
                        elif col_name in ['PriceValue', 'CountSold', 'CountReview']:
                            format_code = '0' # Format Integer
                        elif col_name == 'Rating':
                            format_code = '0.0' # Format Float 1 desimal (sesuaikan jika perlu 0.00 dsb)
                        # Tidak perlu format khusus untuk URL di sini, akan dibuat hyperlink

                        # Terapkan number format jika ada
                        if format_code:
                            for row_idx in range(2, max_row + 1):
                                cell = worksheet[f"{col_letter}{row_idx}"]
                                cell.number_format = format_code

                        # Buat Hyperlink untuk kolom ProductURL
                        if col_name == 'ProductURL':
                            hyperlink_font = Font(color="0000FF", underline="single") # Style biru + underline
                            for row_idx in range(2, max_row + 1):
                                cell = worksheet[f"{col_letter}{row_idx}"]
                                if cell.value and isinstance(cell.value, str) and cell.value.startswith('http'):
                                    cell.hyperlink = cell.value
                                    cell.font = hyperlink_font # Terapkan style
                                    # cell.value = "Link" # Opsional: ganti teks cell jadi "Link"
                                else:
                                     # Jika bukan URL valid, pastikan format text
                                     cell.number_format = '@'


                    # (Opsional) Atur lebar kolom agar lebih rapi
                    for col_idx_0based, col_name in enumerate(df.columns):
                         col_letter = get_column_letter(col_idx_0based + 1)
                         # Logika penyesuaian lebar (bisa disesuaikan)
                         max_length = 0
                         # Ambil panjang header
                         header_length = len(str(worksheet[f"{col_letter}1"].value))
                         max_length = header_length + 2 # Tambahkan padding

                         # Cek panjang data di beberapa baris awal/akhir untuk estimasi
                         for row_idx in range(2, min(max_row + 1, 22)): # Cek 20 baris data
                            cell_value = worksheet[f"{col_letter}{row_idx}"].value
                            if cell_value:
                                max_length = max(max_length, len(str(cell_value)))

                         # Batasi lebar maksimum
                         adjusted_width = min(max_length + 2, 50) # Lebar maks 50
                         if col_name == 'ProductURL': adjusted_width = 30 # Lebar spesifik URL
                         elif col_name == 'ProductName': adjusted_width = 40 # Lebar spesifik Nama

                         worksheet.column_dimensions[col_letter].width = adjusted_width


                excel_data = output.getvalue()

                st.download_button(
                    label="Download Data Excel",
                    data=excel_data,
                    file_name=f"tokopedia_produk_sid_{sid_input}_formatted.xlsx", # Tambahkan formatted di nama file
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.warning("Tidak ada data produk valid yang berhasil diekstrak.")

st.markdown("---")
st.caption("Scraper Tokopedia vX.Y - Harap gunakan secara bertanggung jawab.")
